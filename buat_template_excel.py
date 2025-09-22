import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def buat_template_excel_knn():
    """
    Membuat template Excel dengan perhitungan KNN manual
    """
    wb = Workbook()
    
    # Sheet 1: Data Training
    ws_data = wb.active
    ws_data.title = "Data_Training"
    
    # Header untuk data training
    headers_training = [
        'ID', 'Nama_Tanaman', 'Label', 'Mean_Red', 'Mean_Green', 'Mean_Blue',
        'Std_Red', 'Std_Green', 'Std_Blue', 'Area', 'Perimeter', 
        'Aspect_Ratio', 'Solidity', 'Compactness', 'Brightness', 'Contrast'
    ]
    
    # Sample data training
    sample_data = [
        [1, 'Jahe', 'Herbal', 120, 180, 90, 25, 30, 20, 1500, 200, 0.8, 0.7, 0.6, 110, 45],
        [2, 'Kunyit', 'Herbal', 110, 170, 85, 22, 28, 18, 1400, 190, 0.75, 0.68, 0.58, 105, 42],
        [3, 'Sirih', 'Herbal', 100, 160, 80, 20, 25, 15, 1200, 180, 0.7, 0.65, 0.55, 100, 40],
        [4, 'Lidah_Buaya', 'Herbal', 95, 155, 75, 18, 22, 12, 1100, 170, 0.65, 0.62, 0.52, 95, 38],
        [5, 'Pandan', 'Herbal', 105, 165, 82, 21, 26, 16, 1300, 185, 0.72, 0.66, 0.56, 102, 41],
        [6, 'Tomat', 'Non-Herbal', 140, 200, 100, 35, 40, 30, 2000, 250, 0.9, 0.8, 0.7, 130, 55],
        [7, 'Bayam', 'Non-Herbal', 130, 190, 95, 32, 38, 28, 1800, 240, 0.85, 0.75, 0.65, 125, 52],
        [8, 'Cabai', 'Non-Herbal', 135, 195, 98, 33, 39, 29, 1900, 245, 0.88, 0.78, 0.68, 128, 54],
        [9, 'Terong', 'Non-Herbal', 145, 205, 105, 38, 42, 32, 2100, 260, 0.92, 0.82, 0.72, 135, 58],
        [10, 'Wortel', 'Non-Herbal', 125, 185, 92, 30, 35, 25, 1700, 235, 0.82, 0.72, 0.62, 120, 50]
    ]
    
    # Tulis header
    for i, header in enumerate(headers_training, 1):
        cell = ws_data.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Tulis data
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_data.cell(row=row_idx, column=col_idx, value=value)
    
    # Sheet 2: Input Data Baru
    ws_input = wb.create_sheet("Input_Data_Baru")
    
    input_headers = ['Parameter', 'Nilai']
    input_data = [
        ['Nama_Tanaman', 'Tanaman_Baru'],
        ['Mean_Red', 115],
        ['Mean_Green', 175],
        ['Mean_Blue', 88],
        ['Std_Red', 24],
        ['Std_Green', 29],
        ['Std_Blue', 19],
        ['Area', 1450],
        ['Perimeter', 195],
        ['Aspect_Ratio', 0.77],
        ['Solidity', 0.69],
        ['Compactness', 0.59],
        ['Brightness', 107],
        ['Contrast', 43]
    ]
    
    # Tulis header input
    for i, header in enumerate(input_headers, 1):
        cell = ws_input.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Tulis data input
    for row_idx, (param, nilai) in enumerate(input_data, 2):
        ws_input.cell(row=row_idx, column=1, value=param)
        ws_input.cell(row=row_idx, column=2, value=nilai)
    
    # Sheet 3: Perhitungan Jarak
    ws_jarak = wb.create_sheet("Perhitungan_Jarak")
    
    # Header untuk perhitungan jarak
    jarak_headers = ['ID', 'Nama_Tanaman', 'Label'] + [f'Diff_{i}' for i in range(1, 14)] + ['Jarak_Euclidean']
    
    # Tulis header
    for i, header in enumerate(jarak_headers, 1):
        cell = ws_jarak.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Formula untuk perhitungan jarak
    for row in range(2, 12):  # 10 data training
        # ID, Nama, Label
        ws_jarak.cell(row=row, column=1, value=f"=Data_Training.A{row}")
        ws_jarak.cell(row=row, column=2, value=f"=Data_Training.B{row}")
        ws_jarak.cell(row=row, column=3, value=f"=Data_Training.C{row}")
        
        # Perhitungan selisih untuk setiap fitur
        for col in range(4, 17):  # 13 fitur
            input_col = chr(ord('B') + col - 4)  # B, C, D, ... untuk input
            training_col = chr(ord('D') + col - 4)  # D, E, F, ... untuk training
            formula = f"=(Input_Data_Baru.{input_col}{col-2}-Data_Training.{training_col}{row})^2"
            ws_jarak.cell(row=row, column=col, value=formula)
        
        # Jarak Euclidean
        formula_jarak = f"=SQRT(SUM(D{row}:P{row}))"
        ws_jarak.cell(row=row, column=17, value=formula_jarak)
    
    # Sheet 4: Hasil KNN
    ws_hasil = wb.create_sheet("Hasil_KNN")
    
    hasil_headers = ['K', 'ID_Terdekat', 'Nama_Terdekat', 'Label_Terdekat', 'Jarak', 'Voting_Herbal', 'Voting_Non_Herbal', 'Prediksi_Final']
    
    # Tulis header
    for i, header in enumerate(hasil_headers, 1):
        cell = ws_hasil.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Formula untuk K=3 terdekat
    for k in range(1, 4):
        row = k + 1
        ws_hasil.cell(row=row, column=1, value=k)
        
        # Menggunakan SMALL untuk mencari jarak terkecil ke-k
        ws_hasil.cell(row=row, column=5, value=f"=SMALL(Perhitungan_Jarak.Q:Q,{k+1})")
        
        # INDEX MATCH untuk mencari data yang sesuai
        ws_hasil.cell(row=row, column=2, value=f"=INDEX(Perhitungan_Jarak.A:A,MATCH(E{row},Perhitungan_Jarak.Q:Q,0))")
        ws_hasil.cell(row=row, column=3, value=f"=INDEX(Perhitungan_Jarak.B:B,MATCH(E{row},Perhitungan_Jarak.Q:Q,0))")
        ws_hasil.cell(row=row, column=4, value=f"=INDEX(Perhitungan_Jarak.C:C,MATCH(E{row},Perhitungan_Jarak.Q:Q,0))")
    
    # Voting
    ws_hasil.cell(row=5, column=6, value="=COUNTIF(D2:D4,\"Herbal\")")
    ws_hasil.cell(row=5, column=7, value="=COUNTIF(D2:D4,\"Non-Herbal\")")
    ws_hasil.cell(row=5, column=8, value="=IF(F5>G5,\"Herbal\",\"Non-Herbal\")")
    
    # Sheet 5: Instruksi
    ws_instruksi = wb.create_sheet("Instruksi")
    
    instruksi_text = [
        "INSTRUKSI PENGGUNAAN TEMPLATE KNN",
        "",
        "1. DATA TRAINING:",
        "   - Sheet berisi 10 data sampel tanaman dengan fitur-fiturnya",
        "   - Setiap baris adalah satu tanaman dengan label Herbal/Non-Herbal",
        "",
        "2. INPUT DATA BARU:",
        "   - Masukkan nilai fitur tanaman yang ingin diprediksi di kolom B",
        "   - Ubah nilai sesuai hasil ekstraksi fitur dari gambar",
        "",
        "3. PERHITUNGAN JARAK:",
        "   - Otomatis menghitung selisih kuadrat setiap fitur",
        "   - Menghitung jarak Euclidean ke semua data training",
        "",
        "4. HASIL KNN:",
        "   - Menampilkan K=3 data terdekat",
        "   - Melakukan voting untuk menentukan prediksi final",
        "",
        "5. CARA MENGGUNAKAN:",
        "   a. Ekstrak fitur gambar menggunakan script Python",
        "   b. Masukkan nilai fitur ke sheet Input_Data_Baru",
        "   c. Lihat hasil prediksi di sheet Hasil_KNN",
        "",
        "FORMULA YANG DIGUNAKAN:",
        "- Jarak Euclidean: √(Σ(xi-yi)²)",
        "- Voting: Mayoritas dari K tetangga terdekat",
        "",
        "Catatan: Pastikan data training sudah dinormalisasi jika diperlukan"
    ]
    
    for row, text in enumerate(instruksi_text, 1):
        cell = ws_instruksi.cell(row=row, column=1, value=text)
        if row == 1:
            cell.font = Font(bold=True, size=14)
            cell.fill = PatternFill(start_color="E91E63", end_color="E91E63", fill_type="solid")
        elif text.startswith(("1.", "2.", "3.", "4.", "5.")):
            cell.font = Font(bold=True)
    
    # Simpan file
    nama_file = r"c:\Users\ASUS\OneDrive\Dokumen\FINAL PEMOGRAMAN WEB SESMETER 4\Template_KNN_Manual.xlsx"
    wb.save(nama_file)
    print(f"Template Excel KNN berhasil dibuat: {nama_file}")

if __name__ == "__main__":
    buat_template_excel_knn()